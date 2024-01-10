from PyPDF2 import PdfReader
import re
import pandas as pd
from sklearn.cluster import DBSCAN
import openpyxl
from openpyxl.drawing.image import Image
from GUI import create_file_selector_window, create_output_alert_window
from datetime import datetime
import os
import sys
from output_formating_utilities import write_dict_to_xlsx, resize_and_convert_image, add_ImgByteArr_to_worksheet, scale_cols_to_max_width



################### functions ############################
##### functions for reading pdf #####
def read_full_pdf(pdf_file_path: str) -> str:
    """
    Reads all pages of a pdf and returns contents as a string
    """
    # Open the PDF file
    with open(pdf_file_path, 'rb') as file:
        reader = PdfReader(file)
        # Initialize an empty string to hold all text
        full_text = ""
        # Iterate over all the pages
        for page_num in range(0,len(reader.pages)):
            # Extract text from the page
            page = reader.pages[page_num]
            text = page.extract_text()
            # Add the text to the full text
            full_text += text
        return full_text

def extract_lateral_load_cases(text: str) -> str:
    """
    Uses regex to find lateral load cases in a string and returns them as a list
    """
    pattern = r"Lateral Load = (.*?)'"
    lateral_load_cases = re.findall(pattern, text)
    return [load_case.replace(" ","") for load_case in lateral_load_cases]

def find_sections(text: str) -> list:
    """
    Uses break text to divide long string into sections and returns sections as list
    Sections are checked to see if they contain entries which is row in our pdf table
    """
    # Splitting the text at each occurrence of the break text
    break_text = "Handle"
    sections = re.split(break_text, text)
    sections_w_entries = [section for section in sections if check_for_entries(section)]
    return sections_w_entries

def extract_entries_from_section(text: str) -> list :
    """
    Uses regex to find entries (row of data in pdf text) in text and returns each row as list of tuples
    """
    pattern = r"(\d+) (SW\d+) ([a-zA-Z]+\d+-?[a-zA-Z]*) (\d+\*?) (T[\w_]*\d+) (\d{1,3}(?:,\d{3})* [a-zA-Z]+) (T[\w_]*\d+) (\d{1,3}(?:,\d{3})* [a-zA-Z]+)"
    lines = re.findall(pattern, text) # finds all cpatures groups and creates list of tuples
    return lines


def check_for_entries(text: str) -> bool:
    """
    Uses regex to check for entry (row of data) and returns boolean
    """
    pattern = r"(\d+) (SW\d+) ([a-zA-Z]+\d+-?[a-zA-Z]*) (\d+\*?) (T[\w_]*\d+) (\d{1,3}(?:,\d{3})* [a-zA-Z]+) (T[\w_]*\d+) (\d{1,3}(?:,\d{3})* [a-zA-Z]+)"
    return bool(re.search(pattern,text))

def ensure_df_in_list_same_length(df_list: list) -> bool:
    """
    Checks that each dataframe in list is the same length
    """
    assert all([len(df)==len(df_list[0]) for df in df_list])


##### functions for manipulating dataframes #####
def convert_coord_pair_to_float(coord: str) -> tuple:
    """
    Takes string in following format: [1234.2, -134] and returns each number as a float 1234.1, -134"""
    if not type(coord) == str:
        coord = coord.str
    coord = coord.strip()
    coord = coord.strip("[]")
    x,y = coord.split(",")
    x = float(x)
    y = float(y)
    return x,y

def remove_comma_convert_to_float(x: str) -> float:
    """
    Takes a string with expected format "12,3423 lb" and converts number to float
    Makes unit a global variable for output
    """
    global unit
    number,unit = x.split(" ")  # TO DO create function that takes unti and appends to column title
    number = number.replace(",","")
    number = float(number)
    return number

def label_close_points_with_dbscan(df: pd.DataFrame,
                                    coord_cols: list, 
                                    out_label: str="label", 
                                    eps:float=3
                                    ) -> pd.DataFrame:
    """
    Use sklearn DBSCAN alogrithm to create a label column from x,y coordiante colummns
    
    Inputs:
    df (pandas.DataFrame): dataframe contianing coord columns
    coord_cols (list): name of columns containing coordinates
    out_label (str): name for new label column
    eps (int): DBSCAN algorithm paramter; 
        maximum distance between two samples for one to be considered
        as in the neighborhood of the other. 
    """
    X = df[coord_cols].values
    clustering = DBSCAN(eps=eps, min_samples=2).fit(X)
    df[out_label] = clustering.labels_
    return df

def create_location_key(label_l: str, label_r: str) -> str:
    """
    Applied to a datframe via lambda fucntion; 
    Combines two strings with an "_"
    """
    return f"{str(label_l)}_{str(label_r)}"

def extract_level(diaphragm_label: str) -> int:
    """
    Extract digits from diaphragm string and returns them as integer
    """
    assert isinstance(diaphragm_label,str)
    digit_list = [character for character in diaphragm_label if character.isdigit()]
    level = "".join([digit for digit in digit_list])
    return int(level)

def find_delta_forces(frame: pd.DataFrame, force_cols: list) -> pd.DataFrame:
    """
    This function is intended to be applied to each dataframe that is grouped by load case and shear wall.
    It returns the elta force between the wall at current level and the wall above it.
    The top wall returns Nan (no wall above it) and it is given its on force value as the delta force
    """
    # Sort the group if needed
    frame = frame.sort_values("level")
    # Calculate the difference between consecutive rows for 'left_tension_float'
    for force_col in force_cols:
        frame[f"delta_{force_col}"] = frame[force_col] - frame[force_col].shift(-1)
        frame[f"delta_{force_col}"] = frame[f"delta_{force_col}"].fillna(frame[force_col])
    return frame


##### Output Related Functions #####
def append_date_and_time(prefix: str) -> str:
    """
    Appends current date to the prefix string
    """
    date = datetime.now()
    formated_date = date.strftime("%m-%d-%Y")
    return f"{prefix}_{formated_date}"

# create folder for report using current date and time
def create_output_folder_path(name: str) -> str:
    """
    Finds CWD and returns output folder path inside desktop dir
    """
    output_folder_name = append_date_and_time(name)
    current_folder = os.getcwd() # get Home directory
    output_folder_path = os.path.join(current_folder, output_folder_name)
    return output_folder_path

def resource_path(relative_path: str) -> str:
    """ 
    Get absolute path to resource, works for dev and for PyInstaller 
    Note: sys._MEIPASS is a temporary folder for PyInstaller, necessary to load images folder into bundled app
    """
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

######### LOGING
def log_message(message, filepath):
    f = open(filepath,"a")
    f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}----{message}\n")
    f.close()


####################### Key variables #####################################
eps = 2
logo_path = resource_path(r"images/MSD Full Logo.jpg")
df_start_row = 12 # for xlsx


LSW_input_file_path, LSW_output_file_path,lsw_model_version, job_info = create_file_selector_window()

# create output folder path in CWD
output_folder_path = create_output_folder_path("LSW_HoldownReport")
print(f"- Output Folder: {output_folder_path}")
#create folder for output
if not os.path.exists(output_folder_path): 
    os.makedirs(output_folder_path)
    print(" * Output Folder created *")

log_file_path =os.path.join(output_folder_path, "log.txt")
xlsx_output_path = os.path.join(output_folder_path, "output.xlsx")
log_message("-"*25+"BEGIN ANALYSIS"+"-"*25, log_file_path)
# # ############### Read output pdf and store as DF #################
output_text = read_full_pdf(LSW_output_file_path)

##### parse output file to output dict in folowing format {load_case: [entry: tuple]} #####
load_cases = extract_lateral_load_cases(output_text)
sections = find_sections(output_text)
output_dict = {}

# if assertion fails, there was an error reading the output_pdf

assert len(sections) ==  len(load_cases)
log_message("LSW OUTPUT PDF: NUMBER OF ENTRY SECTIONS EQUAL TO NUMBER PARSED LOAD CASES", log_file_path)
# extract entries from section and pair entries with assocaited load case in output_dict
for i,section in enumerate(sections):
    lines = extract_entries_from_section(section)
    if not load_cases[i] in output_dict.keys():
        output_dict[load_cases[i]] = []
        print(f"added {load_cases[i]} to ouput dict")
    output_dict[load_cases[i]].extend(lines)

output_cols = [
    'handle',
    'name',
    'diaphragm',
    'shearwall_type',
    'LHD',
    'left_tension',
    'RHD',
    'right_tension',
    'load_case'
]

output_dfs = []

# add load case to each tuple entry from extraction and create dataframe for each load case
for load_case, entries in output_dict.items():
    entries_w_load_case = [entry + (load_case,) for entry in entries ]
    df = pd.DataFrame(entries_w_load_case, columns=output_cols)
    output_dfs.append(df)


ensure_df_in_list_same_length(output_dfs)  #after creating datframes, each load case df should have same length

output_df = pd.concat(output_dfs, ignore_index=True)
output_raw_df = output_df #save raw output for report later

qty_wall_in_output = int(output_df.groupby("load_case").size().mean())
log_message(f"LSW OUTPUT PDF: QTY WALLS w/ OUPUT: {qty_wall_in_output}", log_file_path)
log_message("LSW OUTPUT PDF: SUCCESSFULLY PARSED AND LOADED INTO PANDAS DATAFRAME", log_file_path)

output_df["left_tension"] = output_df.apply(
    lambda row: remove_comma_convert_to_float(row["left_tension"]),
    axis=1
    )
output_df["right_tension"] = output_df.apply(
    lambda row: remove_comma_convert_to_float(row["right_tension"]),
    axis=1
    )

############### Read Input Text and Store as Df ###################
input_cols = [
    "handle",
    "diaphragm",
    "name",
    "shearwall_type",
    "left_location",
    "right_location",
    "LHD",
    "RHD"
]

input_df = pd.read_table(LSW_input_file_path, header=None)
# drop last column which is filled with Nan values
input_df = input_df.iloc[:,:-1] 
input_df.columns = input_cols
# save initial input_df for logging purposes
input_raw_df = input_df 
log_message("LSW INPUT TXT: SUCCESSFULLY LOADED INTO PANDAS DATAFRAME", log_file_path)
log_message(f"LSW INPUT TXT: QTY INDIVIDUAL WALLS: {len(input_df)}", log_file_path)


############## create location keys for input df to determine stacked shear walls
# remove only keep necessary columns for input_df for merge
input_df = input_df[["handle","left_location", "right_location", "diaphragm"]]

input_df[["x_left","y_left"]] = input_df.apply(
    lambda row: convert_coord_pair_to_float(row["left_location"]),
    axis=1,
    result_type="expand"
    )
input_df[["x_right","y_right"]] = input_df.apply(
    lambda row: convert_coord_pair_to_float(row["right_location"]),
    axis=1,
    result_type="expand"
    )


#create unqiue lable for similar left and right coords respectively
input_df = label_close_points_with_dbscan(input_df, ['x_left','y_left'], out_label="label_l",eps=2) 
input_df = label_close_points_with_dbscan(input_df, ['x_right','y_right'], out_label="label_r",eps=2) 

input_df["location_key"] = input_df.apply(
    lambda row: create_location_key(row["label_l"], row["label_r"]),
    axis=1
    )



grouped_sw_df = input_df.groupby("location_key", as_index=False)
input_df["level"] = input_df.apply(lambda row: extract_level(row["diaphragm"]), axis=1)

msg = f"DBSCAN APPLIED AND LOCATION KEY CREATED - Total Stacked Walls: {len(grouped_sw_df)}"
log_message(msg, log_file_path)


########################### Merge Coordinate Keys to Output df
# convert handles from numpy to string for merge
output_df.loc[:,["handle"]] = output_df["handle"].astype("str")  
input_df.loc[:,["handle"]] = input_df["handle"].astype("str")  
df = pd.merge(output_df, input_df, on="handle", how="left")
log_message("INPUT DF LEFT MERGED TO OUTPUT DF", log_file_path)

########################### group by load case and shear wall and find delta forces for each side
grouped_df = df.groupby(["load_case","location_key"])

delta_df = grouped_df.apply(find_delta_forces, force_cols=["left_tension","right_tension"])

delta_df = delta_df.reset_index(drop=True) # convert multi-index back to cols

##### Group by 'handle' and calculate the max delta force for each holdown in shearwall #####
delta_grouped_df = delta_df.groupby("handle")

#obtain index of maximum values for left and right HD
delta_left_max_index = delta_grouped_df["delta_left_tension"].idxmax() 
delta_right_max_index = delta_grouped_df["delta_right_tension"].idxmax()

# filter for max delta vlaue based on index and pull releveant columns for output report
walls_max_delta_left_df = delta_df.loc[delta_left_max_index][
    ["handle","name","delta_left_tension","LHD","load_case","left_location"]
    ]
walls_max_delta_left_df.rename(columns={"load_case": "load_case_left"},inplace=True)

walls_max_delta_right_df = delta_df.loc[delta_right_max_index][
    ["handle","delta_right_tension","RHD","load_case","right_location", "location_key", "level"]
    ]
walls_max_delta_right_df.rename(columns={"load_case": "load_case_right"},inplace=True)

assert len(walls_max_delta_left_df)==len(walls_max_delta_right_df) #dfs should be same length
walls_max_delta_df = walls_max_delta_left_df.merge(walls_max_delta_right_df,how="outer", on="handle") 


# groupy by location key and sort by level for output report clarity
walls_max_delta_df = walls_max_delta_df.groupby(
    "location_key",as_index=False
    ).apply(lambda frame: frame.sort_values("level"))
walls_max_delta_df.drop(columns="location_key",inplace=True)

log_message("MAX DELTA FORCES FOUND FOR WALLS", log_file_path)

############################################## create tiedown schedule
# holdowns are not always same on left and right side for each shearwall
# stack the holddown type and tension force for left and rght shearwalls
LHD_df = walls_max_delta_df[["LHD","delta_left_tension"]]
LHD_df.rename(
    columns={"LHD": "HD_type", "delta_left_tension": "tension"},
    inplace=True
    )
RHD_df = walls_max_delta_df[["RHD","delta_right_tension"]]
RHD_df.rename(
    columns={"RHD": "HD_type", "delta_right_tension": "tension"},
    inplace=True
    )
HD_df = pd.concat([LHD_df,RHD_df],axis=0, ignore_index=True) # concatenat along rows

HD_max_df = HD_df.groupby(
    "HD_type", as_index=False
    ).agg({"tension":"max"}).sort_values("tension",ascending=False) 

# add column for ASD (0.7E)
HD_max_df[f"Tension(0.7E_ASD) [{unit}]"] = (0.7*HD_max_df["tension"]).astype(int)
HD_max_df.rename(columns={"tension": f"Tension(1.0E_LRFD) [{unit}]" }, inplace=True)



# ##################### Output Important dfs to different sheets in output.xlsx #######################

with pd.ExcelWriter(xlsx_output_path) as writer:
    HD_max_df.to_excel(writer, sheet_name="Holdown Summary", index=False, startrow=df_start_row) 
    walls_max_delta_df.to_excel(writer, sheet_name='Wall Info', index=False, startrow=df_start_row) 
    output_raw_df.to_excel(writer, sheet_name='LSW output data', index=False)
    input_raw_df.to_excel(writer, sheet_name='LSW input data', index=False)
    
log_message("DATAFRAMES WRITEN TO .XLSX", log_file_path)
######################### add mar logo to holdown summary, Wall info

wb = openpyxl.load_workbook(xlsx_output_path)
ws_list = [wb["Holdown Summary"],wb["Wall Info"]]

## add headers to sheets in worksheet list            
for ws in ws_list:
    img_byte_array = resize_and_convert_image(logo_path, resize_factor=2.5)
    add_ImgByteArr_to_worksheet(img_byte_array,ws, "A1")
    write_dict_to_xlsx(job_info, ws, start_row=1, start_col=6)
    ws["A7"] = "Max Incremental Seismic Tension per Light Shear Wall (LSW) Analysis"
    ws["A7"].font = openpyxl.styles.Font(bold=True)
    ws["B8"] = "LSW Model Version:"
    ws["C8"] = lsw_model_version
    ws["B9"] = "LSW Analysis Output Files:"
    ws["C9"] = f'"{os.path.basename(LSW_input_file_path)}"'
    ws["C10"] = f'"{os.path.basename(LSW_output_file_path)}"'
    scale_cols_to_max_width(ws, start_row=8)

wb.save(xlsx_output_path)

log_message("HEADERS ADDED TO SPECIFIED WORKSHEETS", log_file_path)
#### display success message
create_output_alert_window(output_folder_path)



 